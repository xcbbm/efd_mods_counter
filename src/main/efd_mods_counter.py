# -*- coding: utf-8 -*-
"""
efd_mods_counter_py.py

说明（中文注释，运行时中文通过 Unicode 拼接或直接使用 utf-8）：
  - 访问 Steam 创意工坊页面（Escape From Duckov），抓取并解析 MOD 总数量
  - 将统计结果写入当前工作区下的 excel 文件夹，以“《逃离鸭科夫》-Mods数量统计.xlsx”为文件名，按行追加
  - 同时生成 latest.txt 调试文件，便于 CI 验证
  - 读取上一次记录并对比，生成中文通知文案并尝试发送系统通知（Windows Toast），失败则打印到控制台

依赖：
  - requests
  - openpyxl

用法（手动执行或计划任务调用）:
  python efd_mods_counter_py.py
"""
import os
import re
import sys
import time
import subprocess
from datetime import datetime, timedelta, timezone

# === 修复模块导入路径 ===
import sys
import os
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
# ========================

# 安全导入短信模块
try:
    from src.main.sms_util import send_mod_count_sms
    SMS_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ 短信模块不可用（依赖缺失或路径错误）: {e}", file=sys.stderr)
    SMS_AVAILABLE = False
except Exception as e:
    print(f"⚠️ 短信模块加载失败: {e}", file=sys.stderr)
    SMS_AVAILABLE = False

# 尝试导入第三方库
try:
    import requests
except Exception:
    requests = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
except Exception:
    load_workbook = None

# ========== 基本配置（可按需调整） ==========
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CFG = {
    'WORKSHOP_URL': 'https://steamcommunity.com/app/3167020/workshop/',
    'USER_AGENT': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0 Safari/537.36',
    'TIMEOUT_SEC': 30,
    'OUTPUT_DIR': os.path.join(PROJECT_ROOT, 'excel'),
    'USE_MIRROR': True,
}

# 定义北京时间（UTC+8）
BEIJING_TZ = timezone(timedelta(hours=8))

def get_beijing_now() -> datetime:
    """获取当前北京时间（带时区信息）"""
    return datetime.now(BEIJING_TZ)

def get_date_str(dt: datetime) -> str:
    """将 datetime 转换为 'YYYY/MM/DD' 字符串（强制使用北京时间）"""
    return dt.astimezone(BEIJING_TZ).strftime('%Y/%m/%d')

# -------------------------------
# 工具：通过十六进制 Unicode 码点构造中文字符串
# -------------------------------

def CN(hexes: str) -> str:
    parts = [p for p in hexes.strip().split() if p]
    chars = []
    for h in parts:
        try:
            chars.append(chr(int(h, 16)))
        except Exception:
            chars.append('?')
    return ''.join(chars)

# -------------------------------
# HTTP 请求 & 解析等工具函数（保持不变）
# -------------------------------

def invoke_http(url: str, timeout: int = 30, user_agent: str = None, use_mirror: bool = True) -> str:
    effective = url
    if use_mirror:
        effective = 'https://r.jina.ai/http://' + re.sub(r'^https?://', '', url, flags=re.I)

    headers = {
        'User-Agent': user_agent or CFG['USER_AGENT'],
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9'
    }

    if requests:
        attempts = 0
        while attempts < 2:
            attempts += 1
            try:
                r = requests.get(effective, headers=headers, timeout=timeout)
                r.raise_for_status()
                r.encoding = r.encoding or 'utf-8'
                return r.text
            except Exception:
                time.sleep(2)

    # 回退到 curl
    try:
        tmp = os.path.join(os.environ.get('TEMP', '/tmp'), f"efd_http_{int(time.time())}.tmp")
        cmd = [
            'curl', '-sS', '-L', '--compressed',
            '-A', user_agent or CFG['USER_AGENT'],
            '-H', 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            '-H', 'Accept-Language: en-US,en;q=0.9',
            effective,
            '--output', tmp
        ]
        subprocess.check_call(cmd)
        with open(tmp, 'r', encoding='utf-8', errors='ignore') as f:
            html = f.read()
        try:
            os.remove(tmp)
        except Exception:
            pass
        return html
    except Exception as e:
        raise RuntimeError(f'HTTP request failed: {e}')

def parse_workshop_mod_count(html: str) -> int:
    patterns = [
        r'See\s+all\s+([\d,\.]+)\s+Mods',
        r'Showing\s+\d+(?:-\d+)?\s+of\s+([\d,\.]+)\s+entries',
        r'id=["\']searchResults_total["\']>\s*([\d,\.]+)\s*<'
    ]
    for p in patterns:
        m = re.search(p, html, flags=re.I)
        if m:
            num = re.sub(r'[^\d]', '', m.group(1))
            if num:
                return int(num)
    raise ValueError('Failed to parse MOD total from Workshop page.')

def get_game_name_cn() -> str:
    return CN('9003 79BB 9E2D 79D1 592B')

def get_excel_path() -> str:
    game = get_game_name_cn()
    cn_num_stat = CN('6570 91CF 7EDF 8BA1')
    filename = f"{game}-Mods{cn_num_stat}.xlsx"
    outdir = CFG['OUTPUT_DIR']
    os.makedirs(outdir, exist_ok=True)
    return os.path.join(outdir, filename)

def ensure_excel_row(excel_path: str, date: datetime, count: int):
    if load_workbook is None:
        raise RuntimeError('openpyxl is required to write Excel files. Please pip install openpyxl')

    date_str = get_date_str(date)
    center_align = Alignment(horizontal='center', vertical='center')

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.worksheets[0]

        for col_idx in range(1, 4):
            col_letter = get_column_letter(col_idx)
            try:
                ws.column_dimensions[col_letter].width = max(getattr(ws.column_dimensions[col_letter], 'width', 0) or 0, 50)
            except Exception:
                try:
                    ws.column_dimensions[col_letter].width = 50
                except Exception:
                    pass

        found = False
        for row in range(2, ws.max_row + 1):
            date_cell = ws.cell(row=row, column=1).value
            if date_cell is None:
                continue
            parsed_date = None
            if isinstance(date_cell, datetime):
                parsed_date = date_cell
            else:
                try:
                    parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
                except Exception:
                    try:
                        parsed_date = datetime.fromisoformat(str(date_cell))
                    except Exception:
                        parsed_date = None

            cmp_str = parsed_date.strftime('%Y/%m/%d') if parsed_date else str(date_cell)
            if cmp_str == date_str:
                ws.cell(row=row, column=3).value = int(count)
                for col in range(1, 4):
                    ws.cell(row=row, column=col).alignment = center_align
                found = True
                break

        if not found:
            ws.append([date_str, get_game_name_cn(), int(count)])
            new_row = ws.max_row
            for col in range(1, 4):
                ws.cell(row=new_row, column=col).alignment = center_align

        wb.save(excel_path)
        wb.close()
        return

    # 创建新文件
    wb = Workbook()
    ws = wb.active
    ws.title = 'ModCounts'

    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 50

    ws.append(['Date', 'Game', 'ModCount'])
    for cell in ws[1]:
        try:
            cell.alignment = center_align
            cell.font = Font(bold=True)
        except Exception:
            pass

    ws.append([date_str, get_game_name_cn(), int(count)])
    new_row = ws.max_row
    for col in range(1, 4):
        try:
            ws.cell(row=new_row, column=col).alignment = center_align
        except Exception:
            pass

    wb.save(excel_path)
    wb.close()

def get_yesterday_count(excel_path: str, target_date_str: str = None):
    if load_workbook is None:
        return None
    if not os.path.exists(excel_path):
        return None
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.worksheets[0]

        if target_date_str:
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                if date_cell is None:
                    continue
                parsed_date = None
                if isinstance(date_cell, datetime):
                    parsed_date = date_cell
                else:
                    try:
                        parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
                    except Exception:
                        try:
                            parsed_date = datetime.fromisoformat(str(date_cell))
                        except Exception:
                            parsed_date = None

                cmp_str = parsed_date.strftime('%Y/%m/%d') if parsed_date else str(date_cell)
                if cmp_str == target_date_str:
                    val = ws.cell(row=row, column=3).value
                    wb.close()
                    try:
                        return (parsed_date, int(val))
                    except Exception:
                        return None
            wb.close()
            return None

        # fallback to last row (not used in main anymore)
        max_row = ws.max_row
        if max_row < 2:
            wb.close()
            return None
        val = ws.cell(row=max_row, column=3).value
        date_cell = ws.cell(row=max_row, column=1).value
        wb.close()
        if val is None or val == '':
            return None
        parsed_date = None
        if isinstance(date_cell, datetime):
            parsed_date = date_cell
        else:
            try:
                parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
            except Exception:
                try:
                    parsed_date = datetime.fromisoformat(str(date_cell))
                except Exception:
                    parsed_date = None
        try:
            return (parsed_date, int(val))
        except Exception:
            return None
    except Exception:
        return None

def send_toast(title: str, message: str):
    print(f"{title} - {message}")

def install_deps():
    try:
        req = os.path.join(PROJECT_ROOT, 'requirements.txt')
        if not os.path.exists(req):
            print('requirements.txt 未找到，请手动创建或通过包管理工具安装依赖。')
            return
        print(f'正在使用 {sys.executable} 安装依赖：{req} ...')
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-r', req])
        print('依赖安装完成。')
    except Exception as e:
        print('自动安装依赖失败：', e, file=sys.stderr)
        raise

if '--install-deps' in sys.argv:
    try:
        install_deps()
        sys.exit(0)
    except Exception:
        sys.exit(1)

# -------------------------------
# 主流程（使用北京时间）
# -------------------------------

def main():
    try:
        html = invoke_http(CFG['WORKSHOP_URL'], timeout=CFG['TIMEOUT_SEC'], user_agent=CFG['USER_AGENT'], use_mirror=CFG['USE_MIRROR'])
        count = parse_workshop_mod_count(html)

        # 使用北京时间
        today_dt = get_beijing_now()
        yesterday_dt = today_dt - timedelta(days=1)

        excel_path = get_excel_path()
        yesterday_str = get_date_str(yesterday_dt)
        yinfo = get_yesterday_count(excel_path, yesterday_str)
        ycount = None
        if yinfo is not None:
            _ydate, _ycount = yinfo
            try:
                ycount = int(_ycount)
            except Exception:
                ycount = None

        # 写入今天的记录（北京时间）
        ensure_excel_row(excel_path, today_dt, count)

        # === 新增：写入 latest.txt 调试文件 ===
        debug_path = os.path.join(CFG['OUTPUT_DIR'], 'latest.txt')
        beijing_time_str = get_beijing_now().strftime('%Y-%m-%d %H:%M:%S')
        with open(debug_path, 'w', encoding='utf-8') as f:
            f.write(f"Date: {get_date_str(today_dt)}\n")
            f.write(f"ModCount: {count}\n")
            f.write(f"WrittenAt: {beijing_time_str} (Beijing Time)\n")

        # 构造通知消息
        cn_comma = chr(0xFF0C)
        cn_excl = chr(0xFF01)
        cn_today = f"{today_dt.year}年{today_dt.month}月{today_dt.day}日"
        cn_workshop = CN('521B 610F 5DE2 574A')
        cn_market = CN('5E02 573A')
        cn_of = CN('7684')
        cn_mod = 'Mod'
        cn_total_num_is = CN('603B 6570 91CF 4E3A')
        cn_unit = CN('4E2A')
        cn_yesterday = CN('6BD4 6628 5929')
        cn_more = CN('591A 4E0A 4E86')
        cn_less = CN('51CF 5C11 4E86')

        cn_game = CN('9003 79BB 9E2D 79D1 592B')
        game_quoted = f"\u300A{cn_game}\u300B"

        prefix = f"{cn_today}{cn_comma}{game_quoted}{cn_workshop}{cn_market}{cn_of}{cn_mod}{cn_total_num_is}{count}{cn_unit}"

        if ycount is not None:
            lparen = chr(0xFF08)
            rparen = chr(0xFF09)
            diff = count - ycount
            if diff >= 0:
                msg = f"{prefix}{cn_comma}{cn_yesterday}{yesterday_dt.day}号{lparen}{ycount}{cn_unit}{rparen}{cn_more}{diff}{cn_unit}{cn_excl}"
            else:
                msg = f"{prefix}{cn_comma}{cn_yesterday}{yesterday_dt.day}号{lparen}{ycount}{cn_unit}{rparen}{cn_less}{-diff}{cn_unit}{cn_excl}"
        else:
            msg = f"{prefix}{cn_excl}"

        send_toast('Steam Mod 统计完成', msg)
        print(msg)
                # === 发送短信通知（批量）===
        if SMS_AVAILABLE and ycount is not None:
            increment = count - ycount
            try:
                success_count, total_count = send_mod_count_sms(
                    todaycount=count,
                    yesterdaycount=ycount,
                    increment=increment
                )
                if total_count == 0:
                    print("⏭️ 无有效手机号，跳过短信发送。")
                elif success_count == total_count:
                    print(f"✅ 短信通知已成功发送至 {total_count} 个号码。")
                else:
                    print(f"⚠️ 短信部分成功：{success_count}/{total_count} 个号码接收成功。", file=sys.stderr)
            except Exception as e:
                print(f"❌ 短信发送异常: {e}", file=sys.stderr)
        elif not SMS_AVAILABLE:
            print("⏭️ 跳过短信发送（模块未加载）。")
        else:
            print("⏭️ 无昨日数据，跳过短信发送。")
        return 0
    except Exception as e:
        err = str(e)
        send_toast('Steam Mod 统计失败', err)
        print('Error:', err, file=sys.stderr)
        return 1

if __name__ == '__main__':
    sys.exit(main())