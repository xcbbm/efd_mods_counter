# -*- coding: utf-8 -*-
"""
efd_mods_counter_py.py

说明（中文注释，运行时中文通过 Unicode 拼接或直接使用 utf-8）：
  - 访问 Steam 创意工坊页面（Escape From Duckov），抓取并解析 MOD 总数量
  - 将统计结果写入当前工作区下的 excel 文件夹，以“逃离雅科夫-Mods数量统计.xlsx”为文件名，按行追加
  - 读取上一次记录并对比，生成中文通知文案并尝试发送系统通知（Windows Toast），失败则打印到控制台

依赖：
  - requests
  - openpyxl
  - win10toast (可选，用于 Windows 通知，缺失时回退到控制台)

用法（手动执行或计划任务调用）:
  python efd_mods_counter_py.py

"""
import os
import re
import sys
import time
import subprocess
from datetime import datetime
from datetime import timedelta

# 尝试导入第三方库
try:
    import requests
except Exception:
    requests = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
except Exception:
    load_workbook = None

# 可选通知库
try:
    from win10toast import ToastNotifier
except Exception:
    ToastNotifier = None

# ========== 基本配置（可按需调整） ==========
# 项目根目录（脚本位于 src/main，将向上三层到项目根）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CFG = {
    'WORKSHOP_URL': 'https://steamcommunity.com/app/3167020/workshop/',
    'USER_AGENT': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0 Safari/537.36',
    'TIMEOUT_SEC': 30,
    'OUTPUT_DIR': os.path.join(PROJECT_ROOT, 'excel'),
    'USE_MIRROR': True,   # 受限网络时建议启用：通过 r.jina.ai 镜像读取公开页面
}

# -------------------------------
# 工具：通过十六进制 Unicode 码点构造中文字符串（传入空格分隔的十六进制码点）
# -------------------------------

def CN(hexes: str) -> str:
    parts = [p for p in hexes.strip().split() if p]
    chars = []
    for h in parts:
        try:
            chars.append(chr(int(h, 16)))
        except Exception:
            chars.append('?')
    return ''.join(chars)

# -------------------------------
# 工具：HTTP 请求（优先 requests，失败回退到系统 curl）
# -------------------------------

def invoke_http(url: str, timeout: int = 30, user_agent: str = None, use_mirror: bool = True) -> str:
    effective = url
    if use_mirror:
        # r.jina.ai/http:// + 原始地址去除协议
        effective = 'https://r.jina.ai/http://' + re.sub(r'^https?://', '', url, flags=re.I)

    headers = {
        'User-Agent': user_agent or CFG['USER_AGENT'],
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9'
    }

    # 优先使用 requests
    if requests:
        attempts = 0
        while attempts < 2:
            attempts += 1
            try:
                r = requests.get(effective, headers=headers, timeout=timeout)
                r.raise_for_status()
                # 尝试以 r.encoding 或 utf-8 解码
                r.encoding = r.encoding or 'utf-8'
                return r.text
            except Exception:
                time.sleep(2)
        # 若失败，继续尝试 curl 回退

    # 回退：使用系统 curl.exe
    try:
        tmp = os.path.join(os.environ.get('TEMP', '/tmp'), f"efd_http_{int(time.time())}.tmp")
        cmd = [
            'curl', '-sS', '-L', '--compressed',
            '-A', user_agent or CFG['USER_AGENT'],
            '-H', 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            '-H', 'Accept-Language: en-US,en;q=0.9',
            effective,
            '--output', tmp
        ]
        subprocess.check_call(cmd)
        with open(tmp, 'r', encoding='utf-8', errors='ignore') as f:
            html = f.read()
        try:
            os.remove(tmp)
        except Exception:
            pass
        return html
    except Exception as e:
        raise RuntimeError(f'HTTP request failed: {e}')

# -------------------------------
# 工具：解析 MOD 总数（兼容常见模式）
# -------------------------------

def parse_workshop_mod_count(html: str) -> int:
    patterns = [
        r'See\s+all\s+([\d,\.]+)\s+Mods',
        r'Showing\s+\d+(?:-\d+)?\s+of\s+([\d,\.]+)\s+entries',
        r'id=["\']searchResults_total["\']>\s*([\d,\.]+)\s*<'
    ]
    for p in patterns:
        m = re.search(p, html, flags=re.I)
        if m:
            num = re.sub(r'[^\d]', '', m.group(1))
            if num:
                return int(num)
    raise ValueError('Failed to parse MOD total from Workshop page.')

# -------------------------------
# 获取游戏中文名
# -------------------------------

def get_game_name_cn() -> str:
    # 使用与原 PowerShell 相同的码点（保留原意）
    return CN('9003 79BB 96C5 79D1 592B')

# -------------------------------
# 构造文件名与路径
# -------------------------------

def get_excel_path() -> str:
    game = get_game_name_cn()
    cn_num_stat = CN('6570 91CF 7EDF 8BA1')  # 数量统计
    filename = f"{game}-Mods{cn_num_stat}.xlsx"
    outdir = CFG['OUTPUT_DIR']
    os.makedirs(outdir, exist_ok=True)
    return os.path.join(outdir, filename)

# -------------------------------
# 写入 Excel（openpyxl），若缺少 openpyxl 会抛出异常
# -------------------------------

def ensure_excel_row(excel_path: str, date: datetime, count: int):
    if load_workbook is None:
        raise RuntimeError('openpyxl is required to write Excel files. Please pip install openpyxl')

    date_str = date.strftime('%Y/%m/%d')
    # 中心对齐样式
    center_align = Alignment(horizontal='center', vertical='center')

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.worksheets[0]

        # 确保前三列宽至少为 50（近似像素单位，使用字符宽度单位直接设置为 50）
        for col_idx in range(1, 4):
            col_letter = get_column_letter(col_idx)
            try:
                ws.column_dimensions[col_letter].width = max(getattr(ws.column_dimensions[col_letter], 'width', 0) or 0, 50)
            except Exception:
                try:
                    ws.column_dimensions[col_letter].width = 50
                except Exception:
                    pass

        # 查找是否已有今天的记录，若有则更新第3列（ModCount）；若没有则追加
        found = False
        for row in range(2, ws.max_row + 1):
            date_cell = ws.cell(row=row, column=1).value
            if date_cell is None:
                continue
            parsed_date = None
            if isinstance(date_cell, datetime):
                parsed_date = date_cell
            else:
                try:
                    parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
                except Exception:
                    try:
                        parsed_date = datetime.fromisoformat(str(date_cell))
                    except Exception:
                        parsed_date = None

            cmp_str = parsed_date.strftime('%Y/%m/%d') if parsed_date else str(date_cell)
            if cmp_str == date_str:
                # 更新 ModCount
                ws.cell(row=row, column=3).value = int(count)
                # 设置对齐
                for col in range(1, 4):
                    ws.cell(row=row, column=col).alignment = center_align
                found = True
                break

        if not found:
            ws.append([date_str, get_game_name_cn(), int(count)])
            new_row = ws.max_row
            for col in range(1, 4):
                ws.cell(row=new_row, column=col).alignment = center_align

        wb.save(excel_path)
        wb.close()
        return

    # 文件不存在：创建一个新的 Excel，设置全局格式为居中，前三列列宽设为 50，并写入表头与当天数据
    wb = Workbook()
    ws = wb.active
    ws.title = 'ModCounts'

    # 设置列宽（近似像素值，openpyxl 使用字符宽度单位，直接用 50）
    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 50

    # 添加表头并居中
    ws.append(['Date', 'Game', 'ModCount'])
    for cell in ws[1]:
        try:
            cell.alignment = center_align
        except Exception:
            pass

    # 添加今天的数据，并设置居中
    ws.append([date_str, get_game_name_cn(), int(count)])
    new_row = ws.max_row
    for col in range(1, 4):
        try:
            ws.cell(row=new_row, column=col).alignment = center_align
        except Exception:
            pass

    wb.save(excel_path)
    wb.close()

# -------------------------------
# 读取最后一行的 ModCount（若不存在返回 None）
# -------------------------------

def get_yesterday_count(excel_path: str, target_date_str: str = None):
    """
    在 Excel 第一列按 'YYYY/MM/DD' 日期字符串查找并返回 (parsed_date, count)。
    - 如果给定 target_date_str，则只在表中查找该日期并返回对应的数量；找不到返回 None（不回退到最后一行）。
    - 如果 target_date_str 为 None，则返回最后一条记录（兼容旧行为）。
    """
    if load_workbook is None:
        return None
    if not os.path.exists(excel_path):
        return None
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.worksheets[0]

        # 按日期查找（从第2行开始，跳过表头）
        if target_date_str:
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                if date_cell is None:
                    continue
                parsed_date = None
                if isinstance(date_cell, datetime):
                    parsed_date = date_cell
                else:
                    try:
                        parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
                    except Exception:
                        try:
                            parsed_date = datetime.fromisoformat(str(date_cell))
                        except Exception:
                            parsed_date = None

                cmp_str = parsed_date.strftime('%Y/%m/%d') if parsed_date else str(date_cell)
                if cmp_str == target_date_str:
                    val = ws.cell(row=row, column=3).value
                    wb.close()
                    try:
                        return (parsed_date, int(val))
                    except Exception:
                        return None
            wb.close()
            # 找不到指定日期时，不回退到最后一行，直接返回 None
            return None

        # target_date_str is None: fallback to original last-row behavior
        max_row = ws.max_row
        if max_row < 2:
            wb.close()
            return None
        val = ws.cell(row=max_row, column=3).value
        date_cell = ws.cell(row=max_row, column=1).value
        wb.close()
        if val is None or val == '':
            return None
        parsed_date = None
        if isinstance(date_cell, datetime):
            parsed_date = date_cell
        else:
            try:
                parsed_date = datetime.strptime(str(date_cell), '%Y/%m/%d')
            except Exception:
                try:
                    parsed_date = datetime.fromisoformat(str(date_cell))
                except Exception:
                    parsed_date = None
        try:
            return (parsed_date, int(val))
        except Exception:
            return None
    except Exception:
        return None

# -------------------------------
# 发送系统通知（尝试 win10toast），失败回退到控制台输出
# -------------------------------

def send_toast(title: str, message: str):
    if ToastNotifier:
        try:
            tn = ToastNotifier()
            tn.show_toast(title, message, duration=8, threaded=False)
            return
        except Exception:
            pass
    # 无法显示通知，回退到打印
    print(f"{title} - {message}")

# 可选：自动安装依赖（在开发环境或首次运行时使用）
def install_deps():
    """通过 pip 安装 requirements.txt 中列出的依赖。"""
    try:
        req = os.path.join(PROJECT_ROOT, 'requirements.txt')
        if not os.path.exists(req):
            print('requirements.txt 未找到，请手动创建或通过包管理工具安装依赖。')
            return
        print(f'正在使用 {sys.executable} 安装依赖：{req} ...')
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-r', req])
        print('依赖安装完成。')
    except Exception as e:
        print('自动安装依赖失败：', e, file=sys.stderr)
        raise

# 支持命令行参数 --install-deps 以便在首次部署时方便安装依赖
if '--install-deps' in sys.argv:
    try:
        install_deps()
        sys.exit(0)
    except Exception:
        sys.exit(1)

# -------------------------------
# 主流程
# -------------------------------

def main():
    try:
        html = invoke_http(CFG['WORKSHOP_URL'], timeout=CFG['TIMEOUT_SEC'], user_agent=CFG['USER_AGENT'], use_mirror=CFG['USE_MIRROR'])
        count = parse_workshop_mod_count(html)

        excel_path = get_excel_path()
        # 按日期查找昨日记录（格式 yyyy/MM/dd），只按日期查找且不回退到最后一条
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime('%Y/%m/%d')
        yinfo = get_yesterday_count(excel_path, yesterday_str)
        ycount = None
        if yinfo is not None:
            _ydate, _ycount = yinfo
            # 使用 Excel 中的数量作为昨日数量
            try:
                ycount = int(_ycount)
            except Exception:
                ycount = None

        # 现在写入今天的数据到 Excel
        ensure_excel_row(excel_path, datetime.now(), count)

        # 构造中文通知文案（尽量使用 CN() 构造部分片段）
        cn_comma = chr(0xFF0C)
        cn_excl = chr(0xFF01)
        cn_today = f"{datetime.now().strftime('%Y')}" + chr(0x5E74) + f"{datetime.now().strftime('%m')}" + chr(0x6708) + f"{datetime.now().strftime('%d')}" + chr(0x65E5)
        cn_workshop = CN('521B 610F 5DE5 574A')  # 创意工坊
        cn_market = CN('5E02 573A')              # 市场
        cn_of = CN('7684')                        # 的
        cn_mod = 'Mod'
        cn_total_num_is = CN('603B 6570 91CF 4E3A')  # 总数量为
        cn_unit = CN('4E2A')                       # 个
        cn_yesterday = CN('6BD4 6628 5929')        # 比昨天
        cn_more = CN('591A 4E0A 67B6 4E86')         # 多上架了
        cn_less = CN('51CF 5C11 4E86')             # 减少了

        # 游戏名使用“鸭”字（而非原先的“雅”），并在消息中加上书名号
        cn_game = CN('9003 79BB 9E2D 79D1 592B')  # 逃离鸭科夫
        game_quoted = f"\u300A{cn_game}\u300B"  # 《逃离鸭科夫》

        prefix = f"{cn_today}{cn_comma}{game_quoted}{cn_workshop}{cn_market}{cn_of}{cn_mod}{cn_total_num_is}{count}{cn_unit}"

        # 使用当前日期的前一天作为“昨天”的日
        yesterday_day = (datetime.now() - timedelta(days=1)).day

        if ycount is not None:
            # 使用全角括号展示昨日数量
            lparen = chr(0xFF08)
            rparen = chr(0xFF09)
            diff = count - ycount
            if diff >= 0:
                msg = f"{prefix}{cn_comma}{cn_yesterday}{yesterday_day}号{lparen}{ycount}{cn_unit}{rparen}{cn_more}{diff}{cn_unit}{cn_excl}"
            else:
                msg = f"{prefix}{cn_comma}{cn_yesterday}{yesterday_day}号{lparen}{ycount}{cn_unit}{rparen}{cn_less}{-diff}{cn_unit}{cn_excl}"
        else:
            # 若无法获取昨日数量，则只显示统计结果（不显示比较句）
            msg = f"{prefix}{cn_excl}"

        send_toast('Steam Mod 统计完成', msg)
        print(msg)
        return 0
    except Exception as e:
        err = str(e)
        send_toast('Steam Mod 统计失败', err)
        print('Error:', err, file=sys.stderr)
        return 1

if __name__ == '__main__':
    sys.exit(main())
